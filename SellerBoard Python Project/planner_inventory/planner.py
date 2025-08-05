try:
    import sys
    import traceback
    import os

    sys.path.append(
        "C:\\Program Files\\ZennoLab\\RU\\ZennoPoster Pro V7\\7.7.0.0\\Progs\\Projects\\AMZ Professional\\SellerBoard Python Project\\utilities")
    sys.path.append(
        "D:\\PyCharm Projects\\SellerBoard Python Project\\utilities")

    import time
    import requests
    import re
    import json
    from datetime import datetime, timedelta, timezone
    from dateutil.relativedelta import relativedelta
    import xlrd
    from openpyxl import load_workbook
    from io import BytesIO
    from tables import Table
    from dotenv import load_dotenv
    from log_in import start_log_in
    from sellerboard_interaction import (manage_products_or_planner,
                                         user_agent, headers, read_json_atomic, update_file_atomic, update_local_file)
    load_dotenv()
except Exception as e:
    traceback.print_exc()
    time.sleep(5)
    raise e


file_path = "sellerboard-accounts.txt"
accounts = {}

tracker = []
pages_counter = 1

all_accounts_sheet_name = "Acc"
all_accounts_id_table = os.getenv("ACCOUNTS_SPREADSHEET_ID")
all_accounts_table = Table(all_accounts_id_table)
all_accounts_sheet = all_accounts_table.read_range(all_accounts_sheet_name)


def products(count_switch, session):
    while count_switch != 2:
        dashboard_url = "https://app.sellerboard.com/en/inventory"

        headers = {
            'host': 'app.sellerboard.com',
            'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'upgrade-insecure-requests': '1',
            'user-agent': user_agent,
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-user': '?1',
            'sec-fetch-dest': 'document',
            'referer': 'https://app.sellerboard.com/en/inventory',
            'accept-encoding': 'gzip, deflate, br, zstd',
            'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
            'priority': 'u=0, i',

        }

        dashboard_response = session.get(dashboard_url, headers=headers)

        dashboard_resp = dashboard_response.text

        if dashboard_response.status_code == 200:

            print("Successfully received the dashboard page!")

            csrf_token_match = re.search(r'"CSRF_TOKEN":"(.*?)"', dashboard_resp)
            if csrf_token_match:
                csrf_token = csrf_token_match.group(1)

            csrf_token_key_match = re.search(r'"CSRF_TOKEN_KEY":"(.*?)"', dashboard_resp)
            if csrf_token_key_match:
                csrf_token_key = csrf_token_key_match.group(1)

            print("csrf_token", csrf_token)
            print("csrf_token_key", csrf_token_key)
            print()
            break

    return csrf_token_key, csrf_token, session


def entries_export(session, data_token_key, data_token_value, sb_acc_id, sb_user_id):
    print("Start entriesExport")
    url = "https://app.sellerboard.com/en/inventory/export"

    headers = {
        'User-Agent': user_agent,
        'Sellerboard-Account-Id': sb_acc_id,
        'Sellerboard-User-Id': sb_user_id,
        'X-Requested-With': 'XMLHttpRequest',
        "origin": "https://app.sellerboard.com",
        "pragma": "no-cache",
        "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Opera";v="114"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "accept": "*/*",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "no-cache",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "priority": "u=1, i"
    }

    session.headers.update(headers)

    payload = {
        f'{data_token_key}': data_token_value,

        "type": "stock",

    }

    entries_resp = None

    for i in range(6):

        response = session.post(url, data=payload)

        if response.status_code == 200:

            entries_resp = response.json()

            print(entries_resp)

            if entries_resp['status'] == 'error':
                continue

            # if entries_resp['task_status'] == 'pending':
            #     time.sleep(5)
            #     continue

            return session, entries_resp

        else:
            print(f"{response.status_code} in entriesExport")

            time.sleep(5)

    return session, entries_resp


def export_status(data_token_key, data_token_value, task_id, session):
    print()
    print("Start export status")

    for attempt in range(1, 31):
        print("attempt", attempt)

        url = "https://app.sellerboard.com/en/inventory/exportStatus"

        payload = {
            f'{data_token_key}': data_token_value,
            "taskId": task_id,
            "type": "stock",

        }

        response_post = session.post(url, data=payload, timeout=(60, 120))

        if response_post.status_code == 200:

            try:
                response_data = response_post.json()

                if response_data:

                    redirect = response_data.get("redirect")

                    if redirect:
                        print("redirect", redirect)
                        return True, session
                    else:
                        print("Message from server: Report file is missing")
                        print(response_data)
                        time.sleep(3)
                        continue

                    # print(response_data["report_status"])
                    print(redirect)
                    print()
                    return True, session

                else:
                    print(response_data["report_status"])
                    time.sleep(5)
                    continue
            except requests.exceptions.JSONDecodeError:
                print()
                print("ERROR JSON")
                time.sleep(3)

        else:
            print(response_post.status_code)
            time.sleep(3)
            continue

        time.sleep(3)

    return False, session


def get_data_from_download(json_res, session, sheet_name, spreadsheet_id):
    def fill_table(rows):
        def fill_bonus_tabel_acc_time():

            all_accounts_rows = all_accounts_table.try_table_operation(lambda: all_accounts_sheet.get_all_values())

            planner_update_dates_col = 9
            spreadsheet_id_col = 2

            for row_i, row in enumerate(all_accounts_rows):

                if spreadsheet_id in row[spreadsheet_id_col]:
                    current_date_time = datetime.now().strftime("%d.%m.%Y %H:%M")
                    all_accounts_table.try_table_operation(
                        lambda: all_accounts_sheet.update_cell(row_i + 1, planner_update_dates_col, current_date_time))
                    break

        print("START FILL TABLE")

        t = Table(spreadsheet_id)  # for working
        sheet = t.read_range(sheet_name, spreadsheet_id)

        # Replace all None values in the rows list with an empty string
        all_rows = [[cell if cell is not None else '' for cell in row] for row in rows]

        all_rows.extend(t.empty_last_rows(sheet, all_rows))
        t.update_range(sheet, all_rows, 1)

        fill_bonus_tabel_acc_time()

    def get_filtered_rows():
        headers_added = False
        filtered_rows = []
        work_sheet = workbook.active

        for row in work_sheet.iter_rows(min_row=1, values_only=True):
            if row[4] not in ('', 0, '0', None) and row[5] not in ('', 0, '0', None):
                if row[17] == "Amazon.com" or "Marketplace":  # "Marketplace" only for headers

                    if not headers_added:
                        headers = ["Date Updated", "Time Updated"] + list(row)
                        filtered_rows.append(headers)
                        are_headers_added = True

                    else:
                        current_date = datetime.now().strftime("%d.%m.%Y")
                        current_time = datetime.now().strftime("%H:%M")
                        trimmed_row = [current_date, current_time] + list(row)
                        filtered_rows.append(trimmed_row)
        return filtered_rows

    for i in range(5):

        url = json_res['redirect']
        response_get = session.get(url)

        if response_get.status_code == 200:
            content_type = response_get.headers.get('Content-Type')

            print('Successfully downloaded')

            if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                workbook = load_workbook(filename=BytesIO(response_get.content))

                filtered_rows = get_filtered_rows()

                fill_table(filtered_rows)

                update_local_file(spreadsheet_id)

                return 1


        else:
            print("Status code", response_get.status_code)
            time.sleep(3)


def main():
    base_dir = "../planner_inventory"
    cookies_file = "planner_cookies.json"
    sheet_name = "Inventory"

    manage_products_or_planner(base_dir, cookies_file,
                               products, entries_export, export_status,
                               get_data_from_download, sheet_name
                               )


if __name__ == "__main__":
    big_error = Exception
    try:
        start_time = time.time()
        main()

        end_time = time.time()

        execution_time = end_time - start_time
        print(f"Program execution time: {execution_time:.2f} seconds")

    except Exception as e:
        big_error = e
        print("big_error", str(big_error))
        time.sleep(5)

    if big_error is not Exception:
        raise big_error
